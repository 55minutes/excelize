from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Style


class Book(object):
    def __init__(self, outfile, optimized_write=False):
        self.optimized_write = optimized_write
        self.workbook = Workbook(optimized_write=optimized_write)
        self.outfile = outfile
        self.sheets = []

    def add_sheet(self, sheet):
        sheet.book = self
        if not self.optimized_write and not self.sheets:
            sheet._worksheet = self.workbook.get_active_sheet()
            sheet._worksheet.title = sheet.name
        else:
            sheet._worksheet = self.workbook.create_sheet()
            sheet._worksheet.title = sheet.name
        self.sheets.append(sheet)

    def save(self):
        for sheet in self.sheets:
            sheet.render()
        self.workbook.save(self.outfile)


class Sheet(object):
    def __init__(self, name, rows, title=None, columns=None):
        """
        name: The name of the worksheet.
        rows: And iterable of iterables which constitute the data rows.
        title: A string which is written to cell A1.
        columns: A list of Column objects. These will become column headings.
        """
        for p in ('name', 'rows', 'title', 'columns'):
            setattr(self, p, locals()[p])

    @property
    def worksheet(self):
        return self._worksheet

    def render_title(self):
        if self.title:
            self.worksheet.append([self.title])

    def render_column_headers(self):
        if self.book.optimized_write:
            if self.title:
                self.worksheet.append([""])
            self.worksheet.append([c.name for c in self.columns or []])
        else:
            x = 0
            if self.title:
                x = self.worksheet.get_highest_row() + 2
            for y, column in enumerate(self.columns or [], start=1):
                cell = self.worksheet.cell(row=x, column=y)
                cell.value = column.name
                cell.style = Style(font=Font(bold=True))

    def render_rows(self):
        if self.book.optimized_write:
            for row in self.rows:
                self.worksheet.append(row)
        else:
            for x, row in enumerate(
                    self.rows, start=self.worksheet.get_highest_row() + 1
            ):
                for y, v in enumerate(row, start=1):
                    if self.columns and self.columns[y - 1].is_date:
                        cell = self.worksheet.cell(row=x, column=y)
                        # TODO: We can't assume that the incoming format is
                        #       always going to be an Excel based datetime
                        # Convert from Excel epoch of days since 1900-01-01
                        cell.value = datetime.fromtimestamp(
                            (float(v) - 25569) * 86400
                        )
                    else:
                        cell = self.worksheet.cell(row=x, column=y)
                        cell.value = v

    def render(self):
        self.render_title()
        self.render_column_headers()
        if not self.rows:
            return
        self.render_rows()


class Column(object):
    # TODO: It would be better if we input a custom format, rather than the
    # rigid is_date.
    def __init__(self, name, is_date=False):
        self.name = name
        self.is_date = is_date


def quick_columns(*args):
    cols = []
    for a in args:
        if isinstance(a, list) or isinstance(a, tuple):
            cols.append(Column(a[0], a[1]))
        else:
            cols.append(Column(a))

    return cols
